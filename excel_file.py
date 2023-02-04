from openpyxl import load_workbook
import os
from pyairtable import Table, Api
import time

AIRTABLE_API_KEY = os.getenv("AIRTABLE_API_KEY")
BASE_ID = 'app5geMpic1opZ2nQ'
TABLE_ID = 'tblF32dQfRSfA3zSC'


file = "dividend_file.xlsx"

workbook = load_workbook(filename=file)

sheet = workbook.active
sector_list = []
industry_list = []
headquarter = []
for row in sheet.iter_rows(1, 20, 4, 30, values_only=True):
    print(row)
    #  time.sleep(1)

    # print(f"type {type(row)} row: {row}")
    if not isinstance(row[1], str):
        pass  # print(row[1])
    else:
        industry_list.append(row[1])
    sector_list.append(row[0])

    headquarter.append((row[2]))
    #  dividend_freq

#  print(set(sector_list))
#  print(set(industry_list))
#  print(set(headquarter))


def get_stocks_shares_information(file, min_row, max_row, min_col, max_col, values_only=True):
    # Get requesting stock shares information with specific row and column

    workbook = load_workbook(filename=file)
    sheet = workbook.active
    stock_shares = {}
    for row in sheet.iter_rows(
            min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=values_only):
        try:
            print(f"row14{type(row[14])} : {row[14]}: {str(row[14]/100)}")
        except Exception as error:
            print(f"you have some formula problem: {error}")

        stock_id = row[0]
        try:
            stock = {
                "Ticker": row[0],
                "Company Name": row[1],
                "Sector": row[2],
                "Industry": row[3],
                "Headquarter": row[4],
                "Ex-Div Date": str(row[6]),
                "Div Frequency": row[7],
                "CAPE": row[8],
                "Forward P/E": row[9],
                "P/Projected FCF": row[10],
                "PEG 5-Y": row[11],
                "Cash to Debt": row[12],
                "Median 5-Y ROE %": row[13],
                "FCF Margin %": row[14]/100,
                "Net Margin %": row[15]/100,
                "Forward Div %": row[16],
                "Dividend Payout": row[17],
                "Years of Div. Growth*": row[18]
            }
        except Exception as error:
            print(f"we got some problem: {error}")
        if row[8] == '':
            stock.update(CAPE=0.0)
        elif row[14] == '/':
            stock['FCF Margin %'] = 0.0
        elif row[15] == '/':
            stock['Net Margin %'] = 0.0
        elif row[18] == '':
            print(f"Aaaaaaaaaaaaaaaaaaaaa")

        stock_shares[stock_id] = stock
    return stock_shares


api = Api(AIRTABLE_API_KEY)
table = Table(AIRTABLE_API_KEY, BASE_ID, TABLE_ID)


def update_airtable_database(stocks, airtable_api, airtable_base_id, airtable_table_id):
    for key, value in stock_info.items():
        try:
            table.create(value)
            #  print(f"I have added div freq: {value['Div Frequency']}")
            #  time.sleep(2)
        except Exception as error:
            print(
                f" {value['Ticker']} and div.Growth* {value['Years of Div. Growth*']} and you got error: {error}")
            #  time.sleep(10)


if __name__ == '__main__':

    stock_info = get_stocks_shares_information(
        file="dividend_file.xlsx", min_row=7, max_row=1200, min_col=2, max_col=25)
    # print(stock_info)
    update_airtable_database(stock_info, AIRTABLE_API_KEY, BASE_ID, TABLE_ID)

